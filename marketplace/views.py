from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
from django.utils import timezone
import json
import barcode
from barcode.writer import ImageWriter
import base64
from io import BytesIO
import os
from django.conf import settings
import requests
from PyPDF2 import PdfMerger
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
import decimal

@login_required
def paris_orders(request):
    """Vista para mostrar las órdenes de Paris"""
    status_filter = request.GET.get('status', '')
    processed_filter = request.GET.get('processed', '')
    printed_filter = request.GET.get('printed', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    with connection.cursor() as cursor:
        # Obtener estadísticas de órdenes
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN orden_impresa = 0 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as nuevas,
                COUNT(CASE WHEN orden_impresa = 1 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as por_procesar,
                COUNT(CASE WHEN orden_procesada = 1 AND orden_despachada = 0 THEN 1 END) as por_despachar,
                COUNT(CASE WHEN orden_despachada = 1 THEN 1 END) as despachadas
            FROM paris_orders
            WHERE 1=1
                AND (%s IS NULL OR originOrderDate >= %s)
                AND (%s IS NULL OR originOrderDate <= %s)
        """, [
            date_from if date_from else None,
            date_from if date_from else None,
            date_to if date_to else None,
            date_to if date_to else None
        ])
        
        stats = dict(zip(['nuevas', 'por_procesar', 'por_despachar', 'despachadas'], cursor.fetchone()))

        # Convertir los filtros a los valores correctos
        processed_value = '1' if processed_filter == '1' else '0' if processed_filter == '0' else None
        printed_value = '1' if printed_filter == '1' else '0' if printed_filter == '0' else None
        
        cursor.execute("""
            CALL get_paris_orders(
                %s, %s, %s, %s, %s, %s, %s, %s,
                @p_total_orders, @p_total_amount
            )
        """, [
            status_filter or None,
            processed_value,
            printed_value,
            date_from if date_from else None,
            date_to if date_to else None,
            search_query or None,
            per_page,
            offset
        ])
        
        # Obtener nombres de columnas
        columns = [col[0] for col in cursor.description]
        # Convertir resultados a lista de diccionarios
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Mapear los campos a los nombres que espera la plantilla
        for order in orders:
            order['orden_paris'] = order.get('subOrderNumber', '')
            order['cliente'] = order.get('nombre_cliente', '')
            order['producto'] = order.get('producto_nombre', '')
            order['numero_boleta'] = order.get('numero_boleta', '')
            order['url_boleta'] = order.get('url_boleta', '')
            order['costo_total'] = order.get('costo_total', 0)
            order['fecha_creacion'] = order.get('fecha_creacion', '')
            order['estado_despacho'] = order.get('estado_despacho', '')
            # Asegurarse de que los campos booleanos sean booleanos
            order['orden_impresa'] = bool(order.get('orden_impresa', False))
            order['orden_procesada'] = bool(order.get('orden_procesada', False))
            order['orden_despachada'] = bool(order.get('orden_despachada', False))
            order['boleta_impresa'] = bool(order.get('boleta_impresa', False))
        
        # Obtener los totales
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()

    status_options = [
        {'value': 'NUEVA', 'label': 'Nueva'},
        {'value': 'PROCESADA', 'label': 'Procesada'},
        {'value': 'ENVIADA', 'label': 'Enviada'}
    ]
    
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'status_filter': status_filter,
        'processed_filter': processed_filter,
        'printed_filter': printed_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'status_options': status_options,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'offset': offset,
        'stats': stats
    }
    
    return render(request, 'marketplace/paris_orders.html', context)

@login_required
def ripley_orders(request):
    """Vista para mostrar las órdenes de Ripley"""
    status_filter = request.GET.get('status', '')
    processed_filter = request.GET.get('processed', '')
    printed_filter = request.GET.get('printed', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    with connection.cursor() as cursor:
        # Obtener estadísticas de órdenes
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN orden_impresa = 0 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as nuevas,
                COUNT(CASE WHEN orden_impresa = 1 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as por_procesar,
                COUNT(CASE WHEN orden_procesada = 1 AND orden_despachada = 0 THEN 1 END) as por_despachar,
                COUNT(CASE WHEN orden_despachada = 1 THEN 1 END) as despachadas
            FROM ripley_orders
            WHERE 1=1
                AND (%s IS NULL OR created_date >= %s)
                AND (%s IS NULL OR created_date <= %s)
        """, [
            date_from if date_from else None,
            date_from if date_from else None,
            date_to if date_to else None,
            date_to if date_to else None
        ])
        
        stats = dict(zip(['nuevas', 'por_procesar', 'por_despachar', 'despachadas'], cursor.fetchone()))

        # Convertir fechas vacías a None
        date_from = date_from if date_from else None
        date_to = date_to if date_to else None
        
        # Llamar al procedimiento almacenado
        cursor.execute("""
            CALL get_ripley_orders(
                %s, %s, %s, %s, %s, %s, %s, %s,
                @total_orders, @total_amount
            )
        """, [
            status_filter or None,
            processed_filter or None,
            printed_filter or None,
            date_from,
            date_to,
            search_query or None,
            per_page,
            offset
        ])
        
        # Obtener resultados
        columns = [col[0] for col in cursor.description]
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Obtener totales
        cursor.execute("SELECT @total_orders, @total_amount")
        total_orders, total_amount = cursor.fetchone()

    status_options = [
        {'value': 'NUEVA', 'label': 'Nueva'},
        {'value': 'PROCESADA', 'label': 'Procesada'},
        {'value': 'ENVIADA', 'label': 'Enviada'},
        {'value': 'CANCELADA', 'label': 'Cancelada'}
    ]
    
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'status_filter': status_filter,
        'processed_filter': processed_filter,
        'printed_filter': printed_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'status_options': status_options,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'offset': offset,
        'stats': stats
    }
    
    return render(request, 'marketplace/ripley_orders.html', context)

@login_required
def falabella_orders(request):
    """Vista para mostrar las órdenes de Falabella"""
    status_filter = request.GET.get('status', '')
    processed_filter = request.GET.get('processed', '')
    printed_filter = request.GET.get('printed', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    with connection.cursor() as cursor:
        # Obtener estadísticas de órdenes
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN orden_impresa = 0 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as nuevas,
                COUNT(CASE WHEN orden_impresa = 1 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as por_procesar,
                COUNT(CASE WHEN orden_procesada = 1 AND orden_despachada = 0 THEN 1 END) as por_despachar,
                COUNT(CASE WHEN orden_despachada = 1 THEN 1 END) as despachadas
            FROM falabella_orders
            WHERE 1=1
                AND (%s IS NULL OR created_at >= %s)
                AND (%s IS NULL OR created_at <= %s)
        """, [
            date_from if date_from else None,
            date_from if date_from else None,
            date_to if date_to else None,
            date_to if date_to else None
        ])
        
        stats = dict(zip(['nuevas', 'por_procesar', 'por_despachar', 'despachadas'], cursor.fetchone()))

        # Convertir fechas vacías a None
        date_from = date_from if date_from else None
        date_to = date_to if date_to else None
        
        # Ajustar el filtro impreso: por defecto solo mostrar NO impresas
        if printed_filter == '':
            printed_filter_value = 0
        elif printed_filter == 'SI':
            printed_filter_value = 1
        elif printed_filter == 'NO':
            printed_filter_value = 0
        else:
            printed_filter_value = None
        
        # Llamar al procedimiento almacenado con COLLATE para manejar las collations
        cursor.execute("""
            CALL get_falabella_orders(
                %s, %s, %s, %s, %s, %s, %s, %s,
                @total_orders, @total_amount
            )
        """, [
            status_filter or None,
            processed_filter or None,
            printed_filter_value,
            date_from,
            date_to,
            search_query or None,
            per_page,
            offset
        ])
        
        # Obtener resultados
        columns = [col[0] for col in cursor.description]
        raw_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Agrupar productos por orden
        orders_dict = {}
        for order in raw_orders:
            orden_id = order.get('orden_falabella')
            if orden_id not in orders_dict:
                orders_dict[orden_id] = order.copy()
                orders_dict[orden_id]['productos'] = []
            # Agregar el producto a la lista
            if order.get('producto'):
                orders_dict[orden_id]['productos'].append(order.get('producto'))
        # Convertir la lista de productos a string separado por |
        orders = []
        for orden in orders_dict.values():
            orden['productos'] = ' | '.join(orden['productos']) if orden['productos'] else ''
            orders.append(orden)

        # Obtener totales
        cursor.execute("SELECT @total_orders, @total_amount")
        total_orders, total_amount = cursor.fetchone()

    status_options = [
        {'value': 'NUEVA', 'label': 'Nueva'},
        {'value': 'PROCESADA', 'label': 'Procesada'},
        {'value': 'ENVIADA', 'label': 'Enviada'},
        {'value': 'CANCELADA', 'label': 'Cancelada'}
    ]
    
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'status_filter': status_filter,
        'processed_filter': processed_filter,
        'printed_filter': printed_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'status_options': status_options,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'offset': offset,
        'stats': stats
    }
    
    return render(request, 'marketplace/falabella_orders.html', context)

@login_required
def mercadolibre_orders(request):
    """Vista para mostrar las órdenes de Mercado Libre"""
    status_filter = request.GET.get('status', '')
    processed_filter = request.GET.get('processed', '')
    printed_filter = request.GET.get('printed', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    with connection.cursor() as cursor:
        # Obtener estadísticas de órdenes
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN orden_impresa = 0 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as nuevas,
                COUNT(CASE WHEN orden_impresa = 1 AND orden_procesada = 0 AND orden_despachada = 0 THEN 1 END) as por_procesar,
                COUNT(CASE WHEN orden_procesada = 1 AND orden_despachada = 0 THEN 1 END) as por_despachar,
                COUNT(CASE WHEN orden_despachada = 1 THEN 1 END) as despachadas
            FROM mercadolibre_orders
            WHERE 1=1
                AND (%s IS NULL OR date_created >= %s)
                AND (%s IS NULL OR date_created <= %s)
                AND (%s IS NULL OR seller_id LIKE %s)
        """, [
            date_from if date_from else None,
            date_from if date_from else None,
            date_to if date_to else None,
            date_to if date_to else None,
            search_query if search_query else None,
            f'%{search_query}%' if search_query else None
        ])
        stats = dict(zip(['nuevas', 'por_procesar', 'por_despachar', 'despachadas'], cursor.fetchone()))

        # Convertir los filtros a los valores correctos
        processed_value = '1' if processed_filter == '1' else '0' if processed_filter == '0' else None
        printed_value = '1' if printed_filter == '1' else '0' if printed_filter == '0' else None
        
        cursor.execute("""
            CALL get_mercadolibre_orders(
                %s, %s, %s, %s, %s, %s, %s, %s,
                @p_total_orders, @p_total_amount
            )
        """, [
            status_filter or None,
            processed_value,
            printed_value,
            date_from or None,
            date_to or None,
            search_query or None,
            per_page,
            offset
        ])
        
        # Obtener nombres de columnas
        columns = [col[0] for col in cursor.description]
        # Convertir resultados a lista de diccionarios
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Obtener los totales
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()

    status_options = [
        {'value': 'NUEVA', 'label': 'Nueva'},
        {'value': 'PROCESADA', 'label': 'Procesada'},
        {'value': 'ENVIADA', 'label': 'Enviada'}
    ]
    
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'status_filter': status_filter,
        'processed_filter': processed_filter,
        'printed_filter': printed_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'status_options': status_options,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'offset': offset,
        'stats': stats
    }
    
    print("Contexto enviado al template:")
    print(f"- Total de órdenes en contexto: {len(context['orders'])}")
    
    return render(request, 'marketplace/mercadolibre_orders.html', context)

@login_required
def order_scanning(request):
    return render(request, 'marketplace/order_scanning.html')

@require_http_methods(["POST"])
@csrf_exempt
def print_multiple_boletas(request):
    """Vista para imprimir múltiples boletas en un solo PDF"""
    try:
        data = json.loads(request.body)
        boleta_urls = data.get('boleta_urls', [])
        
        if not boleta_urls:
            return JsonResponse({'error': 'No se proporcionaron URLs de boletas'}, status=400)
        
        # Aquí iría la lógica para combinar los PDFs
        # Por ahora, solo devolvemos la primera URL para pruebas
        return JsonResponse({'pdf_url': boleta_urls[0]})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def print_picking(request):
    """Vista para generar el documento de picking"""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids:
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    print(f"IDs de órdenes recibidos: {order_ids}")
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_paris_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @p_total_orders, @p_total_amount
            )
        """)
        
        # Obtener nombres de columnas
        columns = [col[0] for col in cursor.description]
        print(f"Columnas obtenidas: {columns}")
        
        # Convertir resultados a lista de diccionarios
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        print(f"Total de órdenes obtenidas: {len(all_orders)}")
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if order.get('subOrderNumber') in order_ids]
        print(f"Órdenes filtradas: {len(orders)}")
        
        # Agrupar productos por SKU
        productos_agrupados = {}
        for order in orders:
            sku = order.get('bsale_sku', '') or 'SIN SKU'  # Usar 'SIN SKU' si es None
            if sku not in productos_agrupados:
                productos_agrupados[sku] = {
                    'producto': order.get('producto_nombre', ''),
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean', '') or 'SIN EAN',  # Usar 'SIN EAN' si es None
                    'cantidad_total': 0,
                    'ordenes': []
                }
            
            cantidad = order.get('cantidad', 1)
            productos_agrupados[sku]['cantidad_total'] += cantidad
            productos_agrupados[sku]['ordenes'].append({
                'orden_paris': order.get('subOrderNumber', ''),
                'cantidad': cantidad
            })
        
        # Convertir el diccionario agrupado a lista y ordenar por SKU
        productos = sorted(productos_agrupados.values(), key=lambda x: x['sku_bsale'] or '')
        
        # Obtener información de las órdenes para el resumen
        ordenes_info = {}
        for order in orders:
            orden_id = order.get('subOrderNumber', '')
            if orden_id not in ordenes_info:
                ordenes_info[orden_id] = {
                    'cliente': order.get('nombre_cliente', ''),
                    'direccion': order.get('direccion_envio', ''),
                    'telefono': order.get('telefono', '')
                }
    
    context = {
        'productos': productos,
        'ordenes_info': ordenes_info,
        'fecha_actual': timezone.now()
    }
    return render(request, 'marketplace/picking.html', context)

@login_required
def print_packing(request):
    """Vista para generar el documento de packing"""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids:
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_paris_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @p_total_orders, @p_total_amount
            )
        """)
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        orders = [order for order in all_orders if order.get('subOrderNumber') in order_ids]
        grouped_orders = {}
        for order in orders:
            order_id = order.get('subOrderNumber', '')
            if order_id not in grouped_orders:
                try:
                    barcode_buffer = BytesIO()
                    barcode_instance = barcode.get('code128', order_id, writer=ImageWriter())
                    options = {
                        'module_width': 1.5,    # Reducido de 2 a 1.5
                        'module_height': 30.0,   # Reducido de 50 a 30
                        'font_size': 0,
                        'text_distance': 1,
                        'quiet_zone': 2,
                        'write_text': False,
                    }
                    barcode_instance.write(barcode_buffer, options)
                    barcode_buffer.seek(0)
                    barcode_base64 = base64.b64encode(barcode_buffer.getvalue()).decode('utf-8')
                except Exception as e:
                    print(f"Error generando código de barras para {order_id}: {e}")
                    barcode_base64 = None
                grouped_orders[order_id] = {
                    'orden_paris': order_id,
                    'cliente': order.get('nombre_cliente', ''),
                    'direccion': order.get('direccion_envio', ''),
                    'telefono': order.get('telefono', ''),
                    'costo_total': order.get('costo_total', 0),
                    'barcode': barcode_base64,
                    'productos': []
                }
            grouped_orders[order_id]['productos'].append({
                'producto': order.get('producto_nombre', ''),
                'sku_bsale': order.get('bsale_sku', '') or 'SIN SKU',
                'ean_bsale': order.get('ean', '') or 'SIN EAN',
                'cantidad': order.get('cantidad', 1)
            })
        orders = list(grouped_orders.values())
    context = {
        'orders': orders,
        'fecha_actual': timezone.now()
    }
    return render(request, 'marketplace/packing.html', context)

def ripley_order_detail(request, order_id):
    # Lógica para obtener los detalles de la orden de Ripley
    return render(request, 'marketplace/ripley_order_detail.html', {'order_id': order_id})

def paris_order_detail(request, order_id):
    # Lógica para obtener los detalles de la orden de Paris
    return render(request, 'marketplace/paris_order_detail.html', {'order_id': order_id})

def falabella_order_detail(request, order_id):
    """Vista para mostrar el detalle de una orden de Falabella"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                fo.order_number AS orden_falabella,
                CONCAT(fo.customer_first_name, ' ', fo.customer_last_name) AS cliente,
                fosa.address1 AS calle,
                fosa.city AS ciudad,
                fosa.region AS region,
                foi.name AS producto,
                bd.number AS numero_boleta,
                bd.urlPdf AS url_boleta,
                bdd.variant_code AS sku_bsale,
                bv.barCode AS ean_bsale,
                bd.netAmount AS costo_neto,
                bd.taxAmount AS iva,
                bd.totalAmount AS costo_total,
                fo.shipping_fee_total AS costo_despacho,
                fo.status AS estado_despacho,
                fo.created_at AS fecha_creacion,
                fo.updated_at AS fecha_actualizacion,
                fo.printed,
                fo.processed
            FROM falabella_orders fo
            JOIN falabella_order_items foi ON foi.order_id = fo.order_id
            LEFT JOIN falabella_orders_shipping_address fosa ON fosa.order_id = fo.order_id
            JOIN bsale_references br ON br.number = fo.order_number
            JOIN bsale_documents bd ON bd.id = br.document_id
            JOIN bsale_document_details bdd ON bdd.document_id = bd.id
            JOIN bsale_variants bv ON bv.id = bdd.variant_id
            WHERE fo.order_number = %s
        """, [order_id])
        
        columns = [col[0] for col in cursor.description]
        order = dict(zip(columns, cursor.fetchone() or []))
    
    return render(request, 'marketplace/falabella_order_detail.html', {'order': order})

def mercadolibre_order_detail(request, order_id):
    """Vista para mostrar el detalle de una orden de Mercado Libre"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                mo.order_id AS orden_mercadolibre,
                mo.buyer_nickname AS cliente,
                moi.title AS producto,
                moi.sku AS sku_mercadolibre,
                moi.seller_sku AS sku_vendedor,
                bd.number AS numero_boleta,
                bd.urlPdf AS url_boleta,
                bd.netAmount AS costo_neto,
                bd.taxAmount AS iva,
                bd.totalAmount AS costo_total,
                mo.shipping_cost AS costo_despacho,
                mo.status AS estado_despacho,
                mo.created_at AS fecha_creacion,
                mo.updated_at AS fecha_actualizacion,
                mo.printed,
                mo.processed
            FROM mercadolibre_orders mo
            JOIN mercadolibre_order_items moi ON moi.order_id = mo.order_id
            JOIN bsale_references br ON br.number = mo.order_id
            JOIN bsale_documents bd ON bd.id = br.document_id
            WHERE mo.order_id = %s
        """, [order_id])
        
        columns = [col[0] for col in cursor.description]
        order = dict(zip(columns, cursor.fetchone() or []))
    
    return render(request, 'marketplace/mercadolibre_order_detail.html', {'order': order})

@require_http_methods(["POST"])
@csrf_exempt
def unir_boletas(request):
    """Recibe una lista de URLs de PDFs, los une y devuelve un solo PDF."""
    import json
    data = json.loads(request.body)
    urls = data.get('boleta_urls', [])
    if not urls:
        return JsonResponse({'error': 'No se proporcionaron URLs de boletas'}, status=400)
    
    merger = PdfMerger()
    
    for url in urls:
        try:
            response = requests.get(url)
            response.raise_for_status()
            merger.append(BytesIO(response.content))
        except Exception as e:
            print(f"Error descargando o uniendo PDF: {url} - {e}")
    
    output = BytesIO()
    merger.write(output)
    merger.close()
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="boletas_unidas.pdf"'
    return response

@login_required
def print_ripley_picking(request):
    """Vista para imprimir el picking de órdenes Ripley."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        # Obtener las órdenes usando el procedimiento almacenado
        cursor.execute("""
            CALL get_ripley_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @total_orders, @total_amount
            )
        """)
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if str(order.get('orden_ripley')) in order_ids]
        
        # Crear diccionario para agrupar productos
        productos_agrupados = {}
        ordenes_info = {}
        
        # Procesar cada orden
        for order in orders:
            orden_id = order.get('orden_ripley', '')
            sku = order.get('sku_bsale', '') or 'SIN SKU'
            producto = order.get('producto', '')
            cantidad = int(order.get('cantidad', 1))
            
            # Agregar información de la orden
            if orden_id not in ordenes_info:
                ordenes_info[orden_id] = {
                    'cliente': order.get('cliente', ''),
                    'direccion': order.get('calle', ''),
                    'comuna': order.get('ciudad', ''),
                    'ciudad': order.get('region', ''),
                    'telefono': order.get('telefono', '')
                }
            
            # Agrupar productos
            if sku not in productos_agrupados:
                productos_agrupados[sku] = {
                    'producto': producto,
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean_bsale', '') or 'SIN EAN',
                    'cantidad_total': 0,
                    'ordenes': []
                }
            
            # Agregar la orden al producto
            productos_agrupados[sku]['ordenes'].append({
                'orden_ripley': orden_id,
                'cantidad': cantidad
            })
            productos_agrupados[sku]['cantidad_total'] += cantidad
        
        # Convertir a lista y ordenar por SKU
        productos = sorted(productos_agrupados.values(), key=lambda x: x['sku_bsale'] or '')
        
        print(f"Total de productos diferentes: {len(productos)}")
        if productos:
            print("Primer producto:", productos[0])
    
    return render(request, 'marketplace/ripley_picking.html', {
        'productos': productos,
        'ordenes_info': ordenes_info,
        'fecha_actual': timezone.now()
    })

@login_required
def print_ripley_packing(request):
    """Vista para imprimir el packing de órdenes Ripley."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_ripley_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @total_orders, @total_amount
            )
        """)
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if str(order.get('orden_ripley')) in order_ids]
        
        # Agrupar órdenes
        grouped_orders = {}
        for order in orders:
            order_id = order.get('orden_ripley', '')
            if order_id not in grouped_orders:
                try:
                    barcode_buffer = BytesIO()
                    barcode_instance = barcode.get('code128', order_id, writer=ImageWriter())
                    options = {
                        'module_width': 1.5,    # Reducido de 2 a 1.5
                        'module_height': 30.0,   # Reducido de 50 a 30
                        'font_size': 0,
                        'text_distance': 1,
                        'quiet_zone': 2,
                        'write_text': False,
                    }
                    barcode_instance.write(barcode_buffer, options)
                    barcode_buffer.seek(0)
                    barcode_base64 = base64.b64encode(barcode_buffer.getvalue()).decode('utf-8')
                except Exception as e:
                    print(f"Error generando código de barras para {order_id}: {e}")
                    barcode_base64 = None
                
                grouped_orders[order_id] = {
                    'orden_ripley': order_id,
                    'cliente': order.get('cliente', ''),
                    'direccion': order.get('calle', ''),
                    'comuna': order.get('ciudad', ''),
                    'ciudad': order.get('region', ''),
                    'codigo_postal': order.get('codigo_postal', ''),
                    'telefono': order.get('telefono', ''),
                    'barcode': barcode_base64,
                    'productos': [],
                    'costo_neto': float(order.get('costo_neto', 0)),
                    'iva': float(order.get('iva', 0)),
                    'costo_total': float(order.get('costo_total', 0)),
                    'costo_despacho': float(order.get('costo_despacho', 0))
                }
            
            # Solo agregar el producto si tiene SKU y no es un costo de despacho
            sku = order.get('sku_bsale', '')
            if sku and sku != 'SIN SKU' and sku != 'dispatch-cost-lp':
                producto = {
                    'producto': order.get('producto', ''),
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean_bsale', '') or 'SIN EAN',
                    'cantidad': int(order.get('cantidad', 1)),
                    'precio': float(order.get('costo_neto', 0))
                }
                
                # Verificar si el producto ya existe en la orden
                producto_existente = next((p for p in grouped_orders[order_id]['productos'] if p['sku_bsale'] == producto['sku_bsale']), None)
                if producto_existente:
                    producto_existente['cantidad'] += producto['cantidad']
                else:
                    grouped_orders[order_id]['productos'].append(producto)
        
        orders = list(grouped_orders.values())
        print(f"Órdenes agrupadas: {len(orders)}")
        if orders:
            print("Primera orden agrupada:", orders[0])
    
    return render(request, 'marketplace/ripley_packing.html', {
        'orders': orders,
        'fecha_actual': timezone.now()
    })

@login_required
def print_falabella_picking(request):
    """Vista para imprimir el picking de órdenes Falabella."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_falabella_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @total_orders, @total_amount
            )
        """)
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if str(order.get('orden_falabella')) in order_ids]
        
        # Crear diccionario para agrupar productos
        productos_agrupados = {}
        ordenes_info = {}
        
        # Procesar cada orden
        for order in orders:
            orden_id = order.get('orden_falabella', '')
            sku = order.get('sku_bsale', '') or 'SIN SKU'
            producto = order.get('producto', '')
            cantidad = int(order.get('cantidad', 1))
            
            # Agregar información de la orden
            if orden_id not in ordenes_info:
                ordenes_info[orden_id] = {
                    'cliente': order.get('cliente', ''),
                    'direccion': order.get('calle', ''),
                    'comuna': order.get('ciudad', ''),
                    'ciudad': order.get('region', ''),
                    'telefono': order.get('telefono', '')
                }
            
            # Agrupar productos
            if sku not in productos_agrupados:
                productos_agrupados[sku] = {
                    'producto': producto,
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean_bsale', '') or 'SIN EAN',
                    'cantidad_total': 0,
                    'ordenes': []
                }
            
            # Agregar la orden al producto
            productos_agrupados[sku]['ordenes'].append({
                'orden_falabella': orden_id,
                'cantidad': cantidad
            })
            productos_agrupados[sku]['cantidad_total'] += cantidad
        
        # Convertir a lista y ordenar por SKU
        productos = sorted(productos_agrupados.values(), key=lambda x: x['sku_bsale'] or '')
        
        print(f"Total de productos diferentes: {len(productos)}")
        if productos:
            print("Primer producto:", productos[0])
    
    return render(request, 'marketplace/falabella_picking.html', {
        'productos': productos,
        'ordenes_info': ordenes_info,
        'fecha_actual': timezone.now()
    })

@login_required
def print_falabella_packing(request):
    """Vista para imprimir el packing de órdenes Falabella."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_falabella_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @total_orders, @total_amount
            )
        """)
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if str(order.get('orden_falabella')) in order_ids]
        
        # Agrupar órdenes
        grouped_orders = {}
        for order in orders:
            order_id = order.get('orden_falabella', '')
            if order_id not in grouped_orders:
                try:
                    barcode_buffer = BytesIO()
                    barcode_instance = barcode.get('code128', order_id, writer=ImageWriter())
                    options = {
                        'module_width': 0.8,    # Reducido para hacer el código más compacto
                        'module_height': 20.0,   # Reducido para hacer el código más compacto
                        'font_size': 0,
                        'text_distance': 1,
                        'quiet_zone': 1,
                        'write_text': False,
                    }
                    barcode_instance.write(barcode_buffer, options)
                    barcode_buffer.seek(0)
                    barcode_base64 = base64.b64encode(barcode_buffer.getvalue()).decode('utf-8')
                except Exception as e:
                    print(f"Error generando código de barras para {order_id}: {e}")
                    barcode_base64 = None
                
                grouped_orders[order_id] = {
                    'orden_falabella': order_id,
                    'cliente': order.get('cliente', ''),
                    'direccion': order.get('calle', ''),
                    'comuna': order.get('ciudad', ''),
                    'ciudad': order.get('region', ''),
                    'telefono': order.get('telefono', ''),
                    'barcode': barcode_base64,
                    'productos': [],
                    'costo_neto': float(order.get('costo_neto', 0)),
                    'iva': float(order.get('iva', 0)),
                    'costo_total': float(order.get('costo_total', 0)),
                    'costo_despacho': float(order.get('costo_despacho', 0))
                }
            
            # Solo agregar el producto si tiene SKU y no es un costo de despacho
            sku = order.get('sku_bsale', '')
            if sku and sku != 'SIN SKU' and sku != 'dispatch-cost-lp':
                producto = {
                    'producto': order.get('producto', ''),
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean_bsale', '') or 'SIN EAN',
                    'cantidad': int(order.get('cantidad', 1)),
                    'precio': float(order.get('costo_neto', 0))
                }
                
                # Verificar si el producto ya existe en la orden
                producto_existente = next((p for p in grouped_orders[order_id]['productos'] if p['sku_bsale'] == producto['sku_bsale']), None)
                if producto_existente:
                    producto_existente['cantidad'] += producto['cantidad']
                else:
                    grouped_orders[order_id]['productos'].append(producto)
        
        orders = list(grouped_orders.values())
        print(f"Órdenes agrupadas: {len(orders)}")
        if orders:
            print("Primera orden agrupada:", orders[0])
    
    return render(request, 'marketplace/falabella_packing.html', {
        'orders': orders,
        'fecha_actual': timezone.now()
    })

@login_required
def print_mercadolibre_picking(request):
    """Vista para imprimir el picking de órdenes Mercado Libre."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        cursor.execute("""
            CALL get_mercadolibre_orders(
                NULL, NULL, NULL, NULL, NULL, NULL, 1000, 0,
                @total_orders, @total_amount
            )
        """)
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Filtrar solo las órdenes seleccionadas
        orders = [order for order in all_orders if str(order.get('orden_mercadolibre')) in order_ids]
        
        # Crear diccionario para agrupar productos
        productos_agrupados = {}
        ordenes_info = {}
        
        # Procesar cada orden
        for order in orders:
            orden_id = order.get('orden_mercadolibre', '')
            sku = order.get('sku_bsale', '') or 'SIN SKU'
            producto = order.get('productos', '')
            cantidad = int(order.get('cantidad', 1))
            
            # Agregar información de la orden
            if orden_id not in ordenes_info:
                ordenes_info[orden_id] = {
                    'cliente': order.get('cliente', ''),
                    'direccion': order.get('direccion', ''),
                    'comuna': order.get('comuna', ''),
                    'ciudad': order.get('ciudad', ''),
                    'telefono': order.get('telefono', '')
                }
            
            # Agrupar productos
            if sku not in productos_agrupados:
                productos_agrupados[sku] = {
                    'producto': producto,
                    'sku_bsale': sku,
                    'ean_bsale': order.get('ean_bsale', '') or 'SIN EAN',
                    'cantidad_total': 0,
                    'ordenes': []
                }
            
            # Agregar la orden al producto
            productos_agrupados[sku]['ordenes'].append({
                'orden_mercadolibre': orden_id,
                'cantidad': cantidad
            })
            productos_agrupados[sku]['cantidad_total'] += cantidad
        
        # Convertir a lista y ordenar por SKU
        productos = sorted(productos_agrupados.values(), key=lambda x: x['sku_bsale'] or '')
        
        print(f"Total de productos diferentes: {len(productos)}")
        if productos:
            print("Primer producto:", productos[0])
    
    return render(request, 'marketplace/mercadolibre_picking.html', {
        'productos': productos,
        'ordenes_info': ordenes_info,
        'fecha_actual': timezone.now()
    })

@login_required
def print_mercadolibre_packing(request):
    """Vista para imprimir el packing de órdenes Mercado Libre."""
    order_ids = request.GET.get('orders', '').split(',')
    if not order_ids or order_ids[0] == '':
        return JsonResponse({'error': 'No se proporcionaron órdenes'}, status=400)
    
    # Eliminar duplicados de order_ids
    order_ids = list(dict.fromkeys(order_ids))
    print(f"IDs de órdenes únicos: {order_ids}")
    
    with connection.cursor() as cursor:
        # Usar la consulta SQL directa en lugar del procedimiento almacenado
        cursor.execute("""
            SELECT
                mo.id AS orden_mercadolibre,
                mo.buyer_nickname AS cliente,
                mo.status AS estado_orden,
                mo.paid_amount AS total_pagado,
                mo.total_amount AS total_orden,
                mo.currency_id AS moneda,
                mo.cancel_detail AS detalle_anulacion,
                mo.tags AS etiquetas_meli,

                GROUP_CONCAT(DISTINCT moi.title SEPARATOR ' | ') AS productos,
                GROUP_CONCAT(DISTINCT moi.seller_sku SEPARATOR ' | ') AS sku_meli,
                GROUP_CONCAT(DISTINCT moi.quantity SEPARATOR ' | ') AS cantidades,
                GROUP_CONCAT(DISTINCT moi.unit_price SEPARATOR ' | ') AS precio_unitario,

                GROUP_CONCAT(DISTINCT bdd.variant_code SEPARATOR ' | ') AS sku_bsale,
                GROUP_CONCAT(DISTINCT bv.barCode SEPARATOR ' | ') AS ean_bsale,
                GROUP_CONCAT(DISTINCT bv.description SEPARATOR ' | ') AS descripcion_bsale,

                bd.number AS numero_boleta,
                bd.urlPdf AS url_boleta,
                bd.netAmount AS costo_neto,
                bd.taxAmount AS iva,
                bd.totalAmount AS costo_total,
                bd.emissionDate AS fecha_emision_boleta,
                bd.address AS direccion_cliente,
                bd.city AS ciudad_cliente,

                mo.date_created AS fecha_creacion,
                mo.last_updated AS fecha_actualizacion,

                mo.orden_impresa,
                mo.orden_procesada,
                mo.orden_despachada,
                mo.boleta_impresa
            FROM mercadolibre_orders mo
            LEFT JOIN mercadolibre_order_items moi ON moi.order_id = mo.id
            JOIN bsale_references br ON br.number = mo.id
            JOIN bsale_documents bd ON bd.id = br.document_id
            JOIN bsale_document_details bdd ON bdd.document_id = bd.id
            JOIN bsale_variants bv ON bv.id = bdd.variant_id
            WHERE mo.id IN %s
            GROUP BY mo.id
            ORDER BY mo.date_created DESC
        """, [tuple(order_ids)])
        
        columns = [col[0] for col in cursor.description]
        all_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Agrupar órdenes
        grouped_orders = {}
        for order in all_orders:
            order_id = str(order.get('orden_mercadolibre', ''))
            if order_id not in grouped_orders:
                try:
                    # Generar código de barras
                    barcode_buffer = BytesIO()
                    barcode_instance = barcode.get('code128', order_id, writer=ImageWriter())
                    options = {
                        'module_width': 1.5,    # Aumentado para hacer el código más grande
                        'module_height': 40.0,   # Aumentado para hacer el código más alto
                        'font_size': 0,
                        'text_distance': 1,
                        'quiet_zone': 1,
                        'write_text': False,
                    }
                    barcode_instance.write(barcode_buffer, options)
                    barcode_buffer.seek(0)
                    barcode_base64 = base64.b64encode(barcode_buffer.getvalue()).decode('utf-8')
                    print(f"Código de barras generado para orden {order_id}")
                except Exception as e:
                    print(f"Error generando código de barras para {order_id}: {e}")
                    barcode_base64 = None
                
                # Obtener la dirección completa y separarla
                direccion_completa = order.get('direccion_cliente', '')
                partes_direccion = direccion_completa.split(',')
                
                # Asignar valores por defecto si no existen
                direccion = partes_direccion[0] if partes_direccion else ''
                comuna = partes_direccion[1].strip() if len(partes_direccion) > 1 else ''
                ciudad = order.get('ciudad_cliente', '')  # Usar el campo ciudad_cliente directamente
                
                grouped_orders[order_id] = {
                    'orden_mercadolibre': order_id,
                    'cliente': order.get('cliente', ''),
                    'direccion': direccion,
                    'comuna': comuna,
                    'ciudad': ciudad,
                    'barcode': barcode_base64,
                    'productos': [],
                    'costo_neto': float(order.get('costo_neto', 0)),
                    'iva': float(order.get('iva', 0)),
                    'costo_total': float(order.get('costo_total', 0)),
                    'costo_despacho': float(order.get('costo_despacho', 0))
                }
            
            # Procesar productos
            productos = order.get('productos', '').split(' | ')
            skus = order.get('sku_bsale', '').split(' | ')
            eans = order.get('ean_bsale', '').split(' | ')
            cantidades = order.get('cantidades', '').split(' | ')
            precios = order.get('precio_unitario', '').split(' | ') if order.get('precio_unitario') else []
            # Si no hay precios unitarios, usar costo_neto dividido por cantidad de productos
            if not precios or all(p.strip() == '' for p in precios):
                try:
                    costo_neto_total = float(order.get('costo_neto', 0))
                    n_productos = len(productos)
                    if n_productos > 0:
                        precio_unitario_fallback = costo_neto_total / n_productos
                        precios = [str(precio_unitario_fallback)] * n_productos
                    else:
                        precios = ['0'] * n_productos
                except Exception:
                    precios = ['0'] * len(productos)
            for i in range(len(productos)):
                if i < len(skus):
                    try:
                        cantidad = int(cantidades[i]) if i < len(cantidades) and cantidades[i].strip() else 1
                        precio = safe_float(precios[i]) if i < len(precios) and precios[i].strip() else 0
                    except (ValueError, TypeError):
                        cantidad = 1
                        precio = 0
                    producto = {
                        'producto': productos[i] if i < len(productos) else '',
                        'sku_bsale': skus[i],
                        'ean_bsale': eans[i] if i < len(eans) else 'SIN EAN',
                        'cantidad': cantidad,
                        'precio': precio,
                        'subtotal': precio * cantidad
                    }
                    
                    # Verificar si el producto ya existe en la orden
                    producto_existente = next((p for p in grouped_orders[order_id]['productos'] if p['sku_bsale'] == producto['sku_bsale']), None)
                    if producto_existente:
                        producto_existente['cantidad'] += producto['cantidad']
                    else:
                        grouped_orders[order_id]['productos'].append(producto)
        
        orders = list(grouped_orders.values())
        print(f"Órdenes agrupadas: {len(orders)}")
        if orders:
            print("Primera orden agrupada:", orders[0])
            print("Código de barras de la primera orden:", orders[0].get('barcode') is not None)
    
    return render(request, 'marketplace/mercadolibre_packing.html', {
        'orders': orders,
        'fecha_actual': timezone.now()
    })

@login_required
def scan_order(request):
    """Vista para escanear y buscar órdenes en todos los marketplaces"""
    order_number = request.GET.get('order_number', '')
    if not order_number:
        return JsonResponse({'error': 'No se proporcionó número de orden'}, status=400)
    
    with connection.cursor() as cursor:
        # Buscar en todas las tablas de marketplaces usando los procedimientos almacenados
        cursor.execute("""
            CALL get_paris_orders(
                NULL, NULL, NULL, NULL, NULL, %s, 1, 0,
                @p_total_orders, @p_total_amount
            )
        """, [order_number])
        columns = [col[0] for col in cursor.description]
        paris_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

        cursor.execute("""
            CALL get_ripley_orders(
                NULL, NULL, NULL, NULL, NULL, %s, 1, 0,
                @total_orders, @total_amount
            )
        """, [order_number])
        columns = [col[0] for col in cursor.description]
        ripley_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

        cursor.execute("""
            CALL get_falabella_orders(
                NULL, NULL, NULL, NULL, NULL, %s, 1, 0,
                @total_orders, @total_amount
            )
        """, [order_number])
        columns = [col[0] for col in cursor.description]
        falabella_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

        cursor.execute("""
            CALL get_mercadolibre_orders(
                NULL, NULL, NULL, NULL, NULL, %s, 1, 0,
                @total_orders, @total_amount
            )
        """, [order_number])
        columns = [col[0] for col in cursor.description]
        mercadolibre_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Agrupar resultados por marketplace
        order_info = {}
        
        # Función auxiliar para convertir valores a float de forma segura
        def safe_float(value, default=0.0):
            try:
                return float(value) if value is not None else default
            except (ValueError, TypeError):
                return default
        
        # Función auxiliar para formatear la información de la orden
        def format_order_info(order, marketplace):
            info = {
                'order_number': '',
                'cliente': '',
                'numero_boleta': '',
                'url_boleta': '',
                'costo_neto': 0.0,
                'iva': 0.0,
                'costo_total': 0.0,
                'estado_despacho': '',
                'fecha_creacion': '',
                'orden_impresa': False,
                'orden_procesada': False,
                'orden_despachada': False,
                'productos': []
            }
            
            if marketplace == 'paris':
                order = paris_orders[0]
                info.update({
                    'order_number': order.get('subOrderNumber', ''),
                    'cliente': order.get('nombre_cliente', ''),
                    'numero_boleta': order.get('numero_boleta', ''),
                    'url_boleta': order.get('url_boleta', ''),
                    'costo_neto': safe_float(order.get('costo_neto')),
                    'iva': safe_float(order.get('iva')),
                    'costo_total': safe_float(order.get('costo_total')),
                    'estado_despacho': order.get('estado_despacho', ''),
                    'fecha_creacion': order.get('fecha_creacion', ''),
                    'orden_impresa': bool(order.get('orden_impresa', False)),
                    'orden_procesada': bool(order.get('orden_procesada', False)),
                    'orden_despachada': bool(order.get('orden_despachada', False)),
                    'productos': [{
                        'producto': order.get('producto_nombre', ''),
                        'sku_bsale': order.get('bsale_sku', ''),
                        'ean_bsale': order.get('ean', ''),
                        'cantidad': int(order.get('cantidad', 1)),
                        'precio': safe_float(order.get('costo_neto')),
                        'subtotal': safe_float(order.get('costo_neto')) * int(order.get('cantidad', 1))
                    }]
                })
            elif marketplace == 'ripley':
                productos = []
                for order in ripley_orders:
                    cantidad = int(order.get('cantidad', 1))
                    precio = safe_float(order.get('costo_neto'))
                    productos.append({
                        'producto': order.get('producto', ''),
                        'sku_bsale': order.get('sku_bsale', ''),
                        'ean_bsale': order.get('ean_bsale', ''),
                        'cantidad': cantidad,
                        'precio': precio,
                        'subtotal': precio * cantidad
                    })
                # Usar la primera orden para los datos generales
                base = ripley_orders[0]
                info = {
                    'order_number': base.get('orden_ripley', ''),
                    'cliente': base.get('cliente', ''),
                    'numero_boleta': base.get('numero_boleta', ''),
                    'url_boleta': base.get('url_boleta', ''),
                    'costo_neto': safe_float(base.get('costo_neto')),
                    'iva': safe_float(base.get('iva')),
                    'costo_total': safe_float(base.get('costo_total')),
                    'estado_despacho': base.get('estado_despacho', ''),
                    'fecha_creacion': base.get('fecha_creacion', ''),
                    'orden_impresa': bool(base.get('orden_impresa', False)),
                    'orden_procesada': bool(base.get('orden_procesada', False)),
                    'orden_despachada': bool(base.get('orden_despachada', False)),
                    'productos': productos
                }
            elif marketplace == 'falabella':
                productos = []
                for order in falabella_orders:
                    cantidad = int(order.get('cantidad', 1))
                    precio = safe_float(order.get('costo_neto'))
                    productos.append({
                        'producto': order.get('producto', ''),
                        'sku_bsale': order.get('sku_bsale', ''),
                        'ean_bsale': order.get('ean_bsale', ''),
                        'cantidad': cantidad,
                        'precio': precio,
                        'subtotal': precio * cantidad
                    })
                base = falabella_orders[0]
                info = {
                    'order_number': base.get('orden_falabella', ''),
                    'cliente': base.get('cliente', ''),
                    'numero_boleta': base.get('numero_boleta', ''),
                    'url_boleta': base.get('url_boleta', ''),
                    'costo_neto': safe_float(base.get('costo_neto')),
                    'iva': safe_float(base.get('iva')),
                    'costo_total': safe_float(base.get('costo_total')),
                    'estado_despacho': base.get('estado_despacho', ''),
                    'fecha_creacion': base.get('fecha_creacion', ''),
                    'orden_impresa': bool(base.get('orden_impresa', False)),
                    'orden_procesada': bool(base.get('orden_procesada', False)),
                    'orden_despachada': bool(base.get('orden_despachada', False)),
                    'productos': productos
                }
            elif marketplace == 'mercadolibre':
                base = mercadolibre_orders[0]
                # Procesar productos concatenados
                productos = []
                productos_nombres = base.get('productos', '').split(' | ')
                skus = base.get('sku_bsale', '').split(' | ')
                eans = base.get('ean_bsale', '').split(' | ')
                cantidades = base.get('cantidades', '').split(' | ')
                precios = base.get('precio_unitario', '').split(' | ') if base.get('precio_unitario') else []
                # Si no hay precios unitarios, usar costo_neto dividido por cantidad de productos
                if not precios or all(p.strip() == '' for p in precios):
                    try:
                        costo_neto_total = float(base.get('costo_neto', 0))
                        n_productos = len(productos_nombres)
                        if n_productos > 0:
                            precio_unitario_fallback = costo_neto_total / n_productos
                            precios = [str(precio_unitario_fallback)] * n_productos
                        else:
                            precios = ['0'] * n_productos
                    except Exception:
                        precios = ['0'] * len(productos_nombres)
                for i in range(len(productos_nombres)):
                    if i < len(skus):
                        try:
                            cantidad = int(cantidades[i]) if i < len(cantidades) and cantidades[i].strip() else 1
                            precio = safe_float(precios[i]) if i < len(precios) and precios[i].strip() else 0
                        except (ValueError, TypeError):
                            cantidad = 1
                            precio = 0
                        productos.append({
                            'producto': productos_nombres[i] if i < len(productos_nombres) else '',
                            'sku_bsale': skus[i],
                            'ean_bsale': eans[i] if i < len(eans) else 'SIN EAN',
                            'cantidad': cantidad,
                            'precio': precio,
                            'subtotal': precio * cantidad
                        })
                info = {
                    'order_number': base.get('orden_mercadolibre', ''),
                    'cliente': base.get('cliente', ''),
                    'numero_boleta': base.get('numero_boleta', ''),
                    'url_boleta': base.get('url_boleta', ''),
                    'costo_neto': safe_float(base.get('costo_neto')),
                    'iva': safe_float(base.get('iva')),
                    'costo_total': safe_float(base.get('costo_total')),
                    'estado_despacho': base.get('estado_orden', ''),
                    'fecha_creacion': base.get('fecha_creacion', ''),
                    'orden_impresa': bool(base.get('orden_impresa', False)),
                    'orden_procesada': bool(base.get('orden_procesada', False)),
                    'orden_despachada': bool(base.get('orden_despachada', False)),
                    'productos': productos
                }
            
            return info
        
        # Procesar órdenes de cada marketplace
        if paris_orders:
            print('--- PARIS ORDERS ---')
            print(paris_orders)
            order_info['paris'] = format_order_info(paris_orders[0], 'paris')
        if ripley_orders:
            print('--- RIPLEY ORDERS ---')
            print(ripley_orders)
            productos = []
            for order in ripley_orders:
                productos.append({
                    'producto': order.get('producto', ''),
                    'sku_bsale': order.get('sku_bsale', ''),
                    'ean_bsale': order.get('ean_bsale', ''),
                    'cantidad': 1,
                    'precio': safe_float(order.get('costo_neto'))
                })
            # Usar la primera orden para los datos generales
            base = ripley_orders[0]
            info = {
                'order_number': base.get('orden_ripley', ''),
                'cliente': base.get('cliente', ''),
                'numero_boleta': base.get('numero_boleta', ''),
                'url_boleta': base.get('url_boleta', ''),
                'costo_neto': safe_float(base.get('costo_neto')),
                'iva': safe_float(base.get('iva')),
                'costo_total': safe_float(base.get('costo_total')),
                'estado_despacho': base.get('estado_despacho', ''),
                'fecha_creacion': base.get('fecha_creacion', ''),
                'orden_impresa': bool(base.get('orden_impresa', False)),
                'orden_procesada': bool(base.get('orden_procesada', False)),
                'orden_despachada': bool(base.get('orden_despachada', False)),
                'productos': productos
            }
            print('Info final RIPLEY:', info)
            order_info['ripley'] = info
        if falabella_orders:
            print('--- FALABELLA ORDERS ---')
            print(falabella_orders)
            productos = []
            for order in falabella_orders:
                productos.append({
                    'producto': order.get('producto', ''),
                    'sku_bsale': order.get('sku_bsale', ''),
                    'ean_bsale': order.get('ean_bsale', ''),
                    'cantidad': 1,
                    'precio': safe_float(order.get('costo_neto'))
                })
            base = falabella_orders[0]
            info = {
                'order_number': base.get('orden_falabella', ''),
                'cliente': base.get('cliente', ''),
                'numero_boleta': base.get('numero_boleta', ''),
                'url_boleta': base.get('url_boleta', ''),
                'costo_neto': safe_float(base.get('costo_neto')),
                'iva': safe_float(base.get('iva')),
                'costo_total': safe_float(base.get('costo_total')),
                'estado_despacho': base.get('estado_despacho', ''),
                'fecha_creacion': base.get('fecha_creacion', ''),
                'orden_impresa': bool(base.get('orden_impresa', False)),
                'orden_procesada': bool(base.get('orden_procesada', False)),
                'orden_despachada': bool(base.get('orden_despachada', False)),
                'productos': productos
            }
            print('Info final FALABELLA:', info)
            order_info['falabella'] = info
        if mercadolibre_orders:
            print('--- MERCADO LIBRE ORDERS ---')
            print(mercadolibre_orders)
            base = mercadolibre_orders[0]
            # Procesar productos concatenados
            productos = []
            productos_nombres = base.get('productos', '').split(' | ')
            skus = base.get('sku_bsale', '').split(' | ')
            eans = base.get('ean_bsale', '').split(' | ')
            cantidades = base.get('cantidades', '').split(' | ')
            precios = base.get('precio_unitario', '').split(' | ') if base.get('precio_unitario') else []
            # Si no hay precios unitarios, usar costo_neto dividido por cantidad de productos
            if not precios or all(p.strip() == '' for p in precios):
                try:
                    costo_neto_total = float(base.get('costo_neto', 0))
                    n_productos = len(productos_nombres)
                    if n_productos > 0:
                        precio_unitario_fallback = costo_neto_total / n_productos
                        precios = [str(precio_unitario_fallback)] * n_productos
                    else:
                        precios = ['0'] * n_productos
                except Exception:
                    precios = ['0'] * len(productos_nombres)
            for i in range(len(productos_nombres)):
                if i < len(skus):
                    try:
                        cantidad = int(cantidades[i]) if i < len(cantidades) and cantidades[i].strip() else 1
                        precio = safe_float(precios[i]) if i < len(precios) and precios[i].strip() else 0
                    except (ValueError, TypeError):
                        cantidad = 1
                        precio = 0
                    productos.append({
                        'producto': productos_nombres[i] if i < len(productos_nombres) else '',
                        'sku_bsale': skus[i],
                        'ean_bsale': eans[i] if i < len(eans) else 'SIN EAN',
                        'cantidad': cantidad,
                        'precio': precio,
                        'subtotal': precio * cantidad
                    })
            info = {
                'order_number': base.get('orden_mercadolibre', ''),
                'cliente': base.get('cliente', ''),
                'numero_boleta': base.get('numero_boleta', ''),
                'url_boleta': base.get('url_boleta', ''),
                'costo_neto': safe_float(base.get('costo_neto')),
                'iva': safe_float(base.get('iva')),
                'costo_total': safe_float(base.get('costo_total')),
                'estado_despacho': base.get('estado_orden', ''),
                'fecha_creacion': base.get('fecha_creacion', ''),
                'orden_impresa': bool(base.get('orden_impresa', False)),
                'orden_procesada': bool(base.get('orden_procesada', False)),
                'orden_despachada': bool(base.get('orden_despachada', False)),
                'productos': productos
            }
            print('Info final MERCADO LIBRE:', info)
            order_info['mercadolibre'] = info
        
        if not order_info:
            return JsonResponse({'error': 'Orden no encontrada'}, status=404)
        
        print('--- ORDER INFO FINAL ---')
        print(order_info)
        
        return JsonResponse({
            'success': True,
            'order_info': order_info
        })

@csrf_exempt
@require_http_methods(["POST"])
def procesar_orden(request):
    """Marca la orden como procesada en el marketplace correspondiente"""
    try:
        data = json.loads(request.body)
        order_number = data.get('order_number')
        marketplace = data.get('marketplace')
        if not order_number or not marketplace:
            return JsonResponse({'success': False, 'error': 'Faltan parámetros'}, status=400)
        with connection.cursor() as cursor:
            if marketplace == 'paris':
                cursor.execute("""
                    UPDATE paris_orders SET orden_procesada = 1 WHERE subOrderNumber = %s
                """, [order_number])
            elif marketplace == 'ripley':
                cursor.execute("""
                    UPDATE ripley_orders SET orden_procesada = 1 WHERE order_id = %s OR commercial_id = %s
                """, [order_number, order_number])
            elif marketplace == 'falabella':
                cursor.execute("""
                    UPDATE falabella_orders SET orden_procesada = 1 WHERE order_number = %s
                """, [str(order_number)])
            elif marketplace == 'mercadolibre':
                cursor.execute("""
                    UPDATE mercadolibre_orders SET orden_procesada = 1 WHERE id = %s
                """, [order_number])
            else:
                return JsonResponse({'success': False, 'error': 'Marketplace no soportado'}, status=400)
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        print('Error al procesar la orden:', e)
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500) 

@require_http_methods(["POST"])
@login_required
def mark_orders_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las órdenes en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE paris_orders 
                SET orden_impresa = 1 
                WHERE subOrderNumber IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_boletas_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las boletas en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE paris_orders 
                SET boleta_impresa = 1 
                WHERE subOrderNumber IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_ripley_orders_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las órdenes en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE ripley_orders 
                SET orden_impresa = 1 
                WHERE order_id IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_ripley_boletas_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las boletas en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE ripley_orders 
                SET boleta_impresa = 1 
                WHERE order_id IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_falabella_orders_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las órdenes en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE falabella_orders 
                SET orden_impresa = 1 
                WHERE order_number IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_falabella_boletas_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        # Actualizar el estado de las boletas en la base de datos
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE falabella_orders 
                SET boleta_impresa = 1 
                WHERE order_number IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_mercadolibre_orders_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE mercadolibre_orders 
                SET orden_impresa = 1 
                WHERE id IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
@login_required
def mark_mercadolibre_boletas_printed(request):
    try:
        data = json.loads(request.body)
        orders = data.get('orders', [])
        
        if not orders:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron órdenes'})
        
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE mercadolibre_orders 
                SET boleta_impresa = 1 
                WHERE id IN %s
            """, [tuple(orders)])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def decimal_to_float(obj):
    if isinstance(obj, list):
        return [decimal_to_float(i) for i in obj]
    if isinstance(obj, tuple):
        return tuple(decimal_to_float(i) for i in obj)
    if isinstance(obj, dict):
        return {k: decimal_to_float(v) for k, v in obj.items()}
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    return obj

@login_required
def paris_sales_analysis(request):
    # Filtros
    year = request.GET.get('year', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        cursor.execute('CALL get_paris_top_products()')
        top_products = cursor.fetchall()
        cursor.execute('CALL get_paris_monthly_top_products()')
        monthly_top_products = cursor.fetchall()
        cursor.execute('CALL get_paris_monthly_status_stats()')
        monthly_status_stats = cursor.fetchall()
        cursor.execute('CALL get_paris_monthly_sales()')
        monthly_sales = cursor.fetchall()
        years = sorted(list(set([str(sale[0]) for sale in monthly_sales if sale[0]])), reverse=True)
        cursor.execute(
            "CALL get_paris_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, per_page, offset]
        )
        columns = [col[0] for col in cursor.description]
        paris_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    context = {
        'top_products': top_products,
        'monthly_top_products': monthly_top_products,
        'monthly_status_stats': monthly_status_stats,
        'monthly_sales': monthly_sales,
        'paris_orders': paris_orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'years': years,
        'year_selected': year,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'per_page': per_page,
        'monthly_sales_json': json.dumps(decimal_to_float(monthly_sales)),
        'monthly_status_stats_json': json.dumps(decimal_to_float(monthly_status_stats)),
    }
    return render(request, 'marketplace/paris_sales_analysis.html', context)

@login_required
def exportar_paris_orders_excel(request):
    year = request.GET.get('year', '')
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        cursor.execute(
            "CALL get_paris_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, 100000, 0]
        )
        columns = [col[0] for col in cursor.description]
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes Paris'
    ws.append(columns)
    for order in orders:
        ws.append([order.get(col, '') for col in columns])
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    from tempfile import NamedTemporaryFile
    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="paris_orders_{year or "todos"}.xlsx"'
    return response

@login_required
def print_paris_picking(request):
    """Vista para imprimir el picking de órdenes de Paris."""
    if request.method == 'POST':
        order_numbers = request.POST.getlist('order_numbers[]')
        if not order_numbers:
            messages.error(request, 'No se seleccionaron órdenes para imprimir.')
            return redirect('paris_orders')
        
        # Obtener las órdenes seleccionadas
        cursor = connection.cursor()
        cursor.execute("""
            SELECT DISTINCT
                po.subOrderNumber,
                pi.sku AS paris_sku,
                pi.name AS producto_nombre,
                bdd.variant_code AS bsale_sku,
                bv.barCode AS ean,
                pi.quantity AS cantidad
            FROM paris_orders po
            JOIN paris_items pi ON pi.subOrderNumber = po.subOrderNumber
            LEFT JOIN bsale_references br ON br.number = po.subOrderNumber
            LEFT JOIN bsale_documents bd ON bd.id = br.document_id
            LEFT JOIN bsale_document_details bdd ON bdd.document_id = bd.id
            LEFT JOIN bsale_variants bv ON bv.id = bdd.variant_id
            WHERE po.subOrderNumber IN %s
            ORDER BY po.subOrderNumber, pi.sku
        """, [tuple(order_numbers)])
        
        orders = cursor.fetchall()
        
        if not orders:
            messages.error(request, 'No se encontraron órdenes para imprimir.')
            return redirect('paris_orders')
        
        # Preparar datos para la plantilla
        picking_data = []
        current_order = None
        current_items = []
        
        for order in orders:
            if current_order != order[0]:
                if current_order is not None:
                    picking_data.append({
                        'order_number': current_order,
                        'items': current_items
                    })
                current_order = order[0]
                current_items = []
            
            current_items.append({
                'paris_sku': order[1],
                'product_name': order[2],
                'bsale_sku': order[3],
                'ean': order[4],
                'quantity': order[5]
            })
        
        if current_order is not None:
            picking_data.append({
                'order_number': current_order,
                'items': current_items
            })
        
        return render(request, 'marketplace/print_picking.html', {
            'picking_data': picking_data,
            'marketplace': 'Paris'
        })
    
    return redirect('paris_orders')

@login_required
def print_paris_packing(request):
    """Vista para imprimir el packing de órdenes de Paris."""
    if request.method == 'POST':
        order_numbers = request.POST.getlist('order_numbers[]')
        if not order_numbers:
            messages.error(request, 'No se seleccionaron órdenes para imprimir.')
            return redirect('paris_orders')
        
        # Obtener las órdenes seleccionadas
        cursor = connection.cursor()
        cursor.execute("""
            SELECT DISTINCT
                po.subOrderNumber,
                po.customer_name,
                CONCAT_WS(', ', po.billing_address1, po.billing_address2, po.billing_address3, po.billing_city) AS direccion_envio,
                po.billing_phone,
                pi.sku AS paris_sku,
                pi.name AS producto_nombre,
                pi.quantity AS cantidad
            FROM paris_orders po
            JOIN paris_items pi ON pi.subOrderNumber = po.subOrderNumber
            WHERE po.subOrderNumber IN %s
            ORDER BY po.subOrderNumber, pi.sku
        """, [tuple(order_numbers)])
        
        orders = cursor.fetchall()
        
        if not orders:
            messages.error(request, 'No se encontraron órdenes para imprimir.')
            return redirect('paris_orders')
        
        # Preparar datos para la plantilla
        packing_data = []
        current_order = None
        current_items = []
        current_customer = None
        current_address = None
        current_phone = None
        
        for order in orders:
            if current_order != order[0]:
                if current_order is not None:
                    packing_data.append({
                        'order_number': current_order,
                        'customer_name': current_customer,
                        'shipping_address': current_address,
                        'phone': current_phone,
                        'items': current_items
                    })
                current_order = order[0]
                current_customer = order[1]
                current_address = order[2]
                current_phone = order[3]
                current_items = []
            
            current_items.append({
                'paris_sku': order[4],
                'product_name': order[5],
                'quantity': order[6]
            })
        
        if current_order is not None:
            packing_data.append({
                'order_number': current_order,
                'customer_name': current_customer,
                'shipping_address': current_address,
                'phone': current_phone,
                'items': current_items
            })
        
        return render(request, 'marketplace/print_packing.html', {
            'packing_data': packing_data,
            'marketplace': 'Paris'
        })
    
    return redirect('paris_orders')

@require_http_methods(["POST"])
@login_required
def mark_paris_orders_printed(request):
    """Marca las órdenes de Paris como impresas."""
    order_numbers = request.POST.getlist('order_numbers[]')
    if not order_numbers:
        return JsonResponse({'status': 'error', 'message': 'No se seleccionaron órdenes.'})
    
    try:
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE paris_orders 
            SET orden_impresa = 1 
            WHERE subOrderNumber IN %s
        """, [tuple(order_numbers)])
        
        return JsonResponse({
            'status': 'success',
            'message': f'Se marcaron {len(order_numbers)} órdenes como impresas.'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error al marcar órdenes como impresas: {str(e)}'
        })

@require_http_methods(["POST"])
@login_required
def mark_paris_boletas_printed(request):
    """Marca las boletas de Paris como impresas."""
    order_numbers = request.POST.getlist('order_numbers[]')
    if not order_numbers:
        return JsonResponse({'status': 'error', 'message': 'No se seleccionaron órdenes.'})
    
    try:
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE paris_orders 
            SET boleta_impresa = 1 
            WHERE subOrderNumber IN %s
        """, [tuple(order_numbers)])
        
        return JsonResponse({
            'status': 'success',
            'message': f'Se marcaron {len(order_numbers)} boletas como impresas.'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error al marcar boletas como impresas: {str(e)}'
        })

@login_required
def ripley_sales_analysis(request):
    import decimal
    year = request.GET.get('year', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        cursor.execute('CALL get_ripley_top_products()')
        top_products = cursor.fetchall()
        cursor.execute('CALL get_ripley_monthly_top_products()')
        monthly_top_products = cursor.fetchall()
        cursor.execute('CALL get_ripley_monthly_status_stats()')
        monthly_status_stats = cursor.fetchall()
        cursor.execute('CALL get_ripley_monthly_sales()')
        monthly_sales = cursor.fetchall()
        years = sorted(list(set([str(sale[0]) for sale in monthly_sales if sale[0]])), reverse=True)
        # Traer órdenes paginadas usando el procedimiento almacenado de órdenes ripley
        cursor.execute(
            "CALL get_ripley_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, per_page, offset]
        )
        columns = [col[0] for col in cursor.description]
        ripley_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    def decimal_to_float(obj):
        if isinstance(obj, list):
            return [decimal_to_float(i) for i in obj]
        if isinstance(obj, tuple):
            return tuple(decimal_to_float(i) for i in obj)
        if isinstance(obj, dict):
            return {k: decimal_to_float(v) for k, v in obj.items()}
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return obj
    context = {
        'top_products': top_products,
        'monthly_top_products': monthly_top_products,
        'monthly_status_stats': monthly_status_stats,
        'monthly_sales': monthly_sales,
        'ripley_orders': ripley_orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'years': years,
        'year_selected': year,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'per_page': per_page,
        'monthly_sales_json': json.dumps(decimal_to_float(monthly_sales)),
        'monthly_status_stats_json': json.dumps(decimal_to_float(monthly_status_stats)),
    }
    return render(request, 'marketplace/ripley_sales_analysis.html', context)

@login_required
def exportar_ripley_orders_excel(request):
    year = request.GET.get('year', '')
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        cursor.execute(
            "CALL get_ripley_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, 100000, 0]
        )
        columns = [col[0] for col in cursor.description]
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
    import openpyxl
    from openpyxl.utils import get_column_letter
    from tempfile import NamedTemporaryFile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes Ripley'
    ws.append(columns)
    for order in orders:
        ws.append([order.get(col, '') for col in columns])
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ripley_orders_{year or 'todos'}.xlsx"'
    return response

@login_required
def mercadolibre_sales_analysis(request):
    import decimal
    year = request.GET.get('year', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        # Si tienes procedimientos para Mercado Libre, descomenta y usa:
        try:
            cursor.execute('CALL get_mercadolibre_top_products()')
            top_products = cursor.fetchall()
        except Exception:
            top_products = []
        try:
            cursor.execute('CALL get_mercadolibre_monthly_top_products()')
            monthly_top_products = cursor.fetchall()
        except Exception:
            monthly_top_products = []
        try:
            cursor.execute('CALL get_mercadolibre_monthly_status_stats()')
            monthly_status_stats = cursor.fetchall()
        except Exception:
            monthly_status_stats = []
        try:
            cursor.execute('CALL get_mercadolibre_monthly_sales()')
            monthly_sales = cursor.fetchall()
        except Exception:
            monthly_sales = []
        # Años disponibles
        years = sorted(list(set([str(sale[0]) for sale in monthly_sales if sale and sale[0]])), reverse=True)
        # Lista de órdenes desde el procedimiento almacenado
        cursor.execute(
            "CALL get_mercadolibre_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, per_page, offset]
        )
        columns = [col[0] for col in cursor.description]
        mercadolibre_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    def decimal_to_float(obj):
        if isinstance(obj, list):
            return [decimal_to_float(i) for i in obj]
        if isinstance(obj, tuple):
            return tuple(decimal_to_float(i) for i in obj)
        if isinstance(obj, dict):
            return {k: decimal_to_float(v) for k, v in obj.items()}
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return obj
    context = {
        'top_products': top_products,
        'monthly_top_products': monthly_top_products,
        'monthly_status_stats': monthly_status_stats,
        'monthly_sales': monthly_sales,
        'mercadolibre_orders': mercadolibre_orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'years': years,
        'year_selected': year,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'per_page': per_page,
        'monthly_sales_json': json.dumps(decimal_to_float(monthly_sales)),
        'monthly_status_stats_json': json.dumps(decimal_to_float(monthly_status_stats)),
    }
    return render(request, 'marketplace/mercadolibre_sales_analysis.html', context)

@login_required
def exportar_mercadolibre_orders_excel(request):
    """Vista para exportar las órdenes de Mercado Libre a Excel"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                seller_id,
                id,
                date_created,
                total_amount,
                status
            FROM mercadolibre_orders
            WHERE 1=1
                AND (%s IS NULL OR date_created >= %s)
                AND (%s IS NULL OR date_created <= %s)
                AND (%s IS NULL OR seller_id LIKE %s)
            ORDER BY date_created DESC
        """, [
            date_from if date_from else None,
            date_from if date_from else None,
            date_to if date_to else None,
            date_to if date_to else None,
            search_query if search_query else None,
            f'%{search_query}%' if search_query else None
        ])
        
        columns = [col[0] for col in cursor.description]
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes Mercado Libre"

    # Agregar encabezados
    headers = ['Vendedor', 'ID Orden', 'Fecha', 'Monto Total', 'Estado']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Agregar datos
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order['seller_id'])
        ws.cell(row=row, column=2, value=order['id'])
        ws.cell(row=row, column=3, value=order['date_created'])
        ws.cell(row=row, column=4, value=float(order['total_amount']))
        ws.cell(row=row, column=5, value=order['status'])

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mercadolibre_orders.xlsx'
    
    wb.save(response)
    return response

def safe_float(value, default=0.0):
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default

@login_required
def falabella_sales_analysis(request):
    import decimal
    year = request.GET.get('year', '')
    page = int(request.GET.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        # 1. Ventas Mensuales
        cursor.execute('''
            SELECT
                YEAR(bd.emissionDate) AS anio,
                MONTH(bd.emissionDate) AS mes,
                COUNT(DISTINCT bd.id) AS total_boletas,
                SUM(bd.netAmount) AS venta_neta,
                SUM(bd.taxAmount) AS total_iva,
                SUM(bd.totalAmount) AS venta_total
            FROM falabella_orders fo
            JOIN bsale_references br ON br.number = fo.order_number
            JOIN bsale_documents bd ON bd.id = br.document_id
            WHERE bd.number IS NOT NULL
            GROUP BY anio, mes
            ORDER BY anio DESC, mes DESC
        ''')
        monthly_sales = cursor.fetchall()
        # 2. Estados de Órdenes
        cursor.execute('''
            SELECT
                YEAR(fo.created_at) AS anio,
                MONTH(fo.created_at) AS mes,
                fo.status AS estado,
                COUNT(DISTINCT fo.order_id) AS total_ordenes
            FROM falabella_orders fo
            GROUP BY anio, mes, estado
            ORDER BY anio DESC, mes DESC
        ''')
        monthly_status_stats = cursor.fetchall()
        # 3. Top 20 Productos Históricos
        cursor.execute('''
            SELECT
                bv.barCode AS ean,
                bv.description AS producto,
                bdd.variant_code AS sku_bsale,
                SUM(bdd.quantity) AS cantidad_vendida,
                SUM(bdd.totalAmount) AS total_vendido
            FROM falabella_orders fo
            JOIN bsale_references br ON br.number = fo.order_number
            JOIN bsale_documents bd ON bd.id = br.document_id
            JOIN bsale_document_details bdd ON bdd.document_id = bd.id
            JOIN bsale_variants bv ON bv.id = bdd.variant_id
            WHERE bd.number IS NOT NULL
            GROUP BY producto, sku_bsale, ean
            ORDER BY cantidad_vendida DESC
            LIMIT 20
        ''')
        top_products = cursor.fetchall()
        # 4. Top 20 Productos por Mes
        cursor.execute('''
            SELECT * FROM (
                SELECT
                    YEAR(bd.emissionDate) AS anio,
                    MONTH(bd.emissionDate) AS mes,
                    bv.barCode AS ean,
                    bv.description AS producto,
                    bdd.variant_code AS sku_bsale,
                    SUM(bdd.quantity) AS cantidad_vendida,
                    SUM(bdd.totalAmount) AS total_vendido,
                    ROW_NUMBER() OVER (
                        PARTITION BY YEAR(bd.emissionDate), MONTH(bd.emissionDate)
                        ORDER BY SUM(bdd.quantity) DESC
                    ) AS fila
                FROM falabella_orders fo
                JOIN bsale_references br ON br.number = fo.order_number
                JOIN bsale_documents bd ON bd.id = br.document_id
                JOIN bsale_document_details bdd ON bdd.document_id = bd.id
                JOIN bsale_variants bv ON bv.id = bdd.variant_id
                WHERE bd.number IS NOT NULL
                GROUP BY anio, mes, producto, sku_bsale, ean
            ) AS sub
            WHERE sub.fila <= 20
            ORDER BY anio DESC, mes DESC, fila
        ''')
        monthly_top_products = cursor.fetchall()
        # Años disponibles
        years = sorted(list(set([str(sale[0]) for sale in monthly_sales if sale and sale[0]])), reverse=True)
        # Lista de órdenes desde el procedimiento almacenado
        cursor.execute(
            "CALL get_falabella_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, per_page, offset]
        )
        columns = [col[0] for col in cursor.description]
        falabella_orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
        cursor.execute("SELECT @p_total_orders, @p_total_amount")
        total_orders, total_amount = cursor.fetchone()
    total_pages = (total_orders + per_page - 1) // per_page if total_orders else 1
    has_previous = page > 1
    has_next = page < total_pages
    previous_page = page - 1
    next_page = page + 1
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    def decimal_to_float(obj):
        if isinstance(obj, list):
            return [decimal_to_float(i) for i in obj]
        if isinstance(obj, tuple):
            return tuple(decimal_to_float(i) for i in obj)
        if isinstance(obj, dict):
            return {k: decimal_to_float(v) for k, v in obj.items()}
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return obj
    context = {
        'top_products': top_products,
        'monthly_top_products': monthly_top_products,
        'monthly_status_stats': monthly_status_stats,
        'monthly_sales': monthly_sales,
        'falabella_orders': falabella_orders,
        'total_orders': total_orders,
        'total_amount': total_amount,
        'years': years,
        'year_selected': year,
        'page': page,
        'total_pages': total_pages,
        'has_previous': has_previous,
        'has_next': has_next,
        'previous_page': previous_page,
        'next_page': next_page,
        'page_range': page_range,
        'per_page': per_page,
        'monthly_sales_json': json.dumps(decimal_to_float(monthly_sales)),
        'monthly_status_stats_json': json.dumps(decimal_to_float(monthly_status_stats)),
    }
    return render(request, 'marketplace/falabella_sales_analysis.html', context)

@login_required
def exportar_falabella_orders_excel(request):
    year = request.GET.get('year', '')
    date_from = f"{year}-01-01" if year else None
    date_to = f"{year}-12-31" if year else None
    with connection.cursor() as cursor:
        cursor.execute(
            "CALL get_falabella_orders(%s, %s, %s, %s, %s, %s, %s, %s, @p_total_orders, @p_total_amount)",
            [None, None, None, date_from, date_to, None, 100000, 0]
        )
        columns = [col[0] for col in cursor.description]
        orders = [dict(zip(columns, row)) for row in cursor.fetchall()]
    import openpyxl
    from openpyxl.utils import get_column_letter
    from tempfile import NamedTemporaryFile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes Falabella'
    ws.append(columns)
    for order in orders:
        ws.append([order.get(col, '') for col in columns])
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="falabella_orders_{year or 'todos'}.xlsx"'
    return response

@login_required
def scan_dispatch(request):
    return render(request, 'marketplace/scan_dispatch.html')

@csrf_exempt
@login_required
def ajax_despachar_orden(request):
    import json
    if request.method == 'POST':
        data = json.loads(request.body)
        order_number = data.get('order_number')
        if not order_number:
            return JsonResponse({'success': False, 'error': 'No se recibió el número de orden.'})
        with connection.cursor() as cursor:
            # Buscar y despachar en Paris
            cursor.execute("SELECT COUNT(*) FROM paris_orders WHERE subOrderNumber = %s", [order_number])
            if cursor.fetchone()[0]:
                cursor.execute("UPDATE paris_orders SET orden_despachada = 1 WHERE subOrderNumber = %s", [order_number])
                return JsonResponse({'success': True, 'marketplace': 'Paris'})
            # Buscar y despachar en Ripley
            cursor.execute("SELECT COUNT(*) FROM ripley_orders WHERE order_id = %s", [order_number])
            if cursor.fetchone()[0]:
                cursor.execute("UPDATE ripley_orders SET orden_despachada = 1 WHERE order_id = %s", [order_number])
                return JsonResponse({'success': True, 'marketplace': 'Ripley'})
            # Buscar y despachar en Falabella
            cursor.execute("SELECT COUNT(*) FROM falabella_orders WHERE order_number = %s", [order_number])
            if cursor.fetchone()[0]:
                cursor.execute("UPDATE falabella_orders SET orden_despachada = 1 WHERE order_number = %s", [order_number])
                return JsonResponse({'success': True, 'marketplace': 'Falabella'})
            # Buscar y despachar en Mercado Libre
            cursor.execute("SELECT COUNT(*) FROM mercadolibre_orders WHERE id = %s", [order_number])
            if cursor.fetchone()[0]:
                cursor.execute("UPDATE mercadolibre_orders SET orden_despachada = 1 WHERE id = %s", [order_number])
                return JsonResponse({'success': True, 'marketplace': 'Mercado Libre'})
        return JsonResponse({'success': False, 'error': 'Orden no encontrada en ningún marketplace.'})